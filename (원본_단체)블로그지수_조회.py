import pandas as pd
import tkinter as tk
from tkinter import filedialog
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.filters import FilterColumn
import random  # 파일 상단에 추가
from time import sleep
import sys
import time  # 이 줄을 추가

# GUI 초기화
root = tk.Tk()
root.withdraw()

# 파일 최상단 import문 아래에 추가
def countdown(seconds):
    for i in range(seconds, 0, -1):
        sys.stdout.write(f"\r대기 시간: {i}초")
        sys.stdout.flush()
        sleep(1)
    sys.stdout.write("\r완료!      \n")
    sys.stdout.flush()

# 입력 엑셀 파일 선택
excel_path = filedialog.askopenfilename(title="블로그 URL이 포함된 엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])

# 저장 경로 미리 선택
save_path = filedialog.asksaveasfilename(title="결과를 저장할 엑셀 파일 선택", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx *.xls")])

# 엑셀 파일 불러오기
df = pd.read_excel(excel_path, usecols=[4], header=0)  # E열에서 데이터 읽기, 첫 행은 헤더

# Selenium 웹드라이버 설정
options = uc.ChromeOptions()
options.add_argument("--disable-gpu")
options.add_argument("--window-size=1920,1080")
options.add_argument('--disable-infobars')
options.add_argument('--lang=ko_KR')

driver = uc.Chrome(options=options)
driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

# Blogdex 로그인 페이지로 이동
url = "https://blogdex.space/login?from=/blog-index"
driver.get(url)
time.sleep(3.2)  # 로그인 이후 타임슬립 3초

# 카카오톡 로그인 버튼 클릭
driver.find_element(By.CLASS_NAME, "border-primary").click()
time.sleep(0.3)
driver.find_element(By.XPATH, "//button[contains(., '카카오톡')]").click()
time.sleep(3.6)  # 로그인 이후 타임슬립 3초

# 로그인 정보 입력
driver.find_element(By.CLASS_NAME, "tf_g").send_keys("01048460380")
time.sleep(1)
password_field = driver.find_elements(By.CLASS_NAME, "tf_g")[1]
password_field.send_keys('@gusqls2')
time.sleep(0.3)
driver.find_element(By.CSS_SELECTOR, ".btn_g.highlight.submit").click()

# 로그인 승인 대기 및 클릭
try:
    login_button = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.CSS_SELECTOR, "button.btn_agree[name='user_oauth_approval'][value='true']")))
    login_button.click()
except Exception as e:
    print("로그인 버튼을 찾을 수 없거나 클릭할 수 없습니다:", e)

# 로그인 후 대기 시간을 10~60초 랜덤으로 설정
wait_time = random.randint(5, 10)
print(f"\n로그인 후 {wait_time}초 대기합니다...")
countdown(wait_time)



data_list = []
save_count = 0
RANDOM_BREAK_COUNT = random.randint(50, 80)  # 50~80개 사이에서 랜덤하게 휴식 시점 설정

for index, row in df.iterrows():
    try:
        blog_url = row[0]
        blog_id = blog_url.split("/")[-1]  # URL에서 블로그 ID 추출
        url_2 = f"https://blogdex.space/blog-index/{blog_id}"
        
        driver.get(url_2)
        sleep_time = random.uniform(8, 10)  # 10~15초로 증가
        time.sleep(sleep_time)  # 아이디마다 크롤링할 때 랜덤 타임슬립

        # 필요한 항목들 스크래핑
        data = {
            "블로그명": driver.find_element(By.CSS_SELECTOR, "div.flex.space-x-1.pr-0.md\\:pr-24 p.text-sm.font-medium.leading-none").text,
            "블로그주소": f"https://blog.naver.com/{blog_id}",
            "블로그지수": driver.find_element(By.CSS_SELECTOR, "svg > text[font-family='Pretendard'][font-size='22px'][font-weight='700'][y='-60']").text,
            "블로그지수(확인용)": driver.find_element(By.CSS_SELECTOR, "div.flex.flex-1.justify-center.px-5 text:nth-child(2)").text,
            "블로그생성일": driver.find_element(By.CSS_SELECTOR, "div.w-full.space-x-1.pr-0.pt-4.md\\:pr-24 div.flex.items-center.justify-center.space-x-2.md\\:justify-start p.text-sm.font-medium.leading-none").text,
            "총방문자": driver.find_element(By.CSS_SELECTOR, "div.w-full.space-x-1.pb-8.pr-0.pt-4.md\\:pb-0.md\\:pr-24 div.flex.items-center.justify-center.space-x-2.md\\:justify-start p.text-sm.font-medium.leading-none").text,
            "총포스팅": driver.find_element(By.CSS_SELECTOR, "#__next > div.flex.min-h-screen.flex-col > main > div > div.flex.flex-col.gap-4 > div:nth-child(1) > div.p-6.pt-0 > div:nth-child(3) > div:nth-child(1) > div > div").text,
            "총구독자": driver.find_element(By.CSS_SELECTOR, "#__next > div.flex.min-h-screen.flex-col > main > div > div.flex.flex-col.gap-4 > div:nth-child(1) > div.p-6.pt-0 > div:nth-child(5) > div:nth-child(1) > div > div").text,
            "주제지수": driver.find_element(By.CSS_SELECTOR, "#__next > div.flex.min-h-screen.flex-col > main > div > div.flex.flex-col.gap-4 > div:nth-child(1) > div.p-6.pt-0 > div.flex.flex-col.justify-center.space-y-12.py-5.md\\:flex-row.md\\:justify-between.md\\:space-x-0.md\\:space-y-0.md\\:py-0 > div.divide.md\\:auto.flex.w-full.flex-1.flex-col.items-center.space-y-4.divide-y.px-5.text-center.md\\:items-end.md\\:text-right > div.pl-0.pt-8.md\\:pl-24.md\\:pt-0 > div > div > p").text,
            "종합지수": driver.find_element(By.CSS_SELECTOR, "#__next > div.flex.min-h-screen.flex-col > main > div > div.flex.flex-col.gap-4 > div:nth-child(1) > div.p-6.pt-0 > div.flex.flex-col.justify-center.space-y-12.py-5.md\\:flex-row.md\\:justify-between.md\\:space-x-0.md\\:space-y-0.md\\:py-0 > div.divide.md\\:auto.flex.w-full.flex-1.flex-col.items-center.space-y-4.divide-y.px-5.text-center.md\\:items-end.md\\:text-right > div:nth-child(2) > div > div > p").text,
            "최고지수": driver.find_element(By.CSS_SELECTOR, "#__next > div.flex.min-h-screen.flex-col > main > div > div.flex.flex-col.gap-4 > div:nth-child(1) > div.p-6.pt-0 > div.flex.flex-col.justify-center.space-y-12.py-5.md\\:flex-row.md\\:justify-between.md\\:space-x-0.md\\:space-y-0.md\\:py-0 > div.divide.md\\:auto.flex.w-full.flex-1.flex-col.items-center.space-y-4.divide-y.px-5.text-center.md\\:items-end.md\\:text-right > div:nth-child(3) > div > div > p").text,
            "블로그주제": driver.find_element(By.CSS_SELECTOR, "div.w-full.pt-4.md\\:w-auto.md\\:pt-0 div.flex.items-center.justify-center.space-x-2.md\\:justify-end p.text-sm.font-medium.leading-none").text,
            "블덱스전체랭킹": driver.find_element(By.CSS_SELECTOR, "#__next > div.flex.min-h-screen.flex-col > main > div > div.flex.flex-col.gap-4 > div:nth-child(1) > div.p-6.pt-0 > div:nth-child(5) > div:nth-child(3) > div > div > p").text,
            "블덱스주제랭킹": driver.find_element(By.CSS_SELECTOR, "div.ml-0.w-full.pt-4.md\\:ml-16.md\\:w-auto.md\\:pt-0 div.flex.items-center.space-x-2.justify-center.md\\:justify-center p.text-sm.font-medium.leading-none").text,
            "최적화수치": driver.find_element(By.CSS_SELECTOR, "div.relative.flex.rounded-md.w-9\\/10 div.bg-primary.h-6.rounded-l-md p.absolute.left-1\\/2.top-1\\/2").text,
            "메일주소": f"{blog_id}@naver.com"
        }
        data_list.append(data)
        save_count += 1

        # 실시간으로 출력
        print(f"아이디: {blog_id}")
        for key, value in data.items():
            print(f"{key}: {value}")
        print("\n")

        # 랜덤한 개수마다 긴 휴식 시간 가지기
        if save_count % RANDOM_BREAK_COUNT == 0:
            long_break = random.randint(180, 300)  # 3~5분 사이 랜덤 휴식
            print(f"\n{save_count}개 수집 완료. {long_break}초 동안 휴식합니다...")
            countdown(long_break)
            RANDOM_BREAK_COUNT = random.randint(50, 80)  # 다음 휴식 시점도 랜덤하게 재설정

        # 50개 단위로 중간 저장
        if save_count % 50 == 0:
            temp_df = pd.DataFrame(data_list)
            temp_save_path = f"{save_path.rsplit('.', 1)[0]}_temp_{save_count}.xlsx"
            temp_df.to_excel(temp_save_path, index=False)
            print(f"{save_count}개의 데이터가 {temp_save_path}에 중간 저장되었습니다.")

    except Exception as e:
        print(f"일부 요소를 찾을 수 없습니다: {e}")
        continue

# DataFrame으로 변환
result_df = pd.DataFrame(data_list)

# 최종 엑셀 파일로 저장
result_df.to_excel(save_path, index=False)

# 엑셀 파일 열기 및 서식 지정
wb = load_workbook(save_path)
ws = wb.active

# 열 너비 설정
column_widths = [18, 35, 12, 12, 14, 10, 9, 9, 9, 9, 9, 12, 20, 20, 12, 24]
for i, width in enumerate(column_widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = width

# '최적' 포함된 셀 서식 변경
red_fill = PatternFill(start_color='FD1E19', end_color='FD1E19', fill_type='solid')
bold_font = Font(bold=True, color='FFFFFF')

for cell in ws['C']:
    if '최적' in str(cell.value):
        cell.font = bold_font
        cell.fill = red_fill

ws.auto_filter.ref = ws.dimensions

# 변경사항 저장
wb.save(save_path)

# 드라이버 종료
driver.quit()

print(f"모든 데이터가 {save_path}에 저장되었습니다.")
